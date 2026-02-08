"""
Script de ingesta: carga datos desde Excel o CSV a PostgreSQL.

Uso:
    python scripts/cargar_datos.py data/raw/excel.xlsx
    python scripts/cargar_datos.py data/raw/excel.xlsx --dry-run
    python scripts/cargar_datos.py --from-csv
"""

import sys
import os
import argparse
import logging
import re

# Agregar directorio raiz al path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pandas as pd

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


def parse_period_to_date(period_str):
    """Convierte periodo string a (anio, trimestre, fecha)."""
    trimestre_mes = {'1\u00ba Trim': (1, 2), '2\u00ba Trim': (2, 5), '3\u00ba Trim': (3, 8), '4\u00ba Trim': (4, 11)}
    parts = period_str.strip().split(' ')
    trim_key = f"{parts[0]} {parts[1]}"
    anio = int(parts[2])
    trimestre, mes = trimestre_mes.get(trim_key, (1, 2))
    from datetime import date
    return anio, trimestre, date(anio, mes, 1)


def load_csv_data():
    """Carga datos desde CSV procesados."""
    data_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data', 'processed')
    data = {}
    for key in ['C1.1', 'C1.2', 'C2.1', 'C2.2', 'C3', 'C4', 'C5', 'C6', 'C7', 'descriptores_CIIU']:
        path = os.path.join(data_dir, f'{key}.csv')
        if os.path.exists(path):
            data[key] = pd.read_csv(path)
            logger.info(f"  {key}: {len(data[key])} filas")
        else:
            logger.warning(f"  {key}: archivo no encontrado")
    return data


def load_excel_data(excel_path):
    """Carga datos desde Excel usando funciones de preprocesamiento."""
    preprocesamiento_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                                          'src', 'preprocesamiento.py')

    # Importar funciones de preprocesamiento
    import importlib.util
    spec = importlib.util.spec_from_file_location("preprocesamiento", preprocesamiento_path)
    prep = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(prep)

    import openpyxl
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    data = {}
    descriptores_all = []

    # Series temporales
    for sheet_name in ['C1.1', 'C1.2', 'C2.1', 'C2.2']:
        if sheet_name in wb.sheetnames:
            data[sheet_name] = prep.process_time_series(wb[sheet_name], sheet_name)
            logger.info(f"  {sheet_name}: {len(data[sheet_name])} filas")

    # Sectoriales
    for sheet_name in ['C3', 'C4', 'C6', 'C7']:
        if sheet_name in wb.sheetnames:
            df_datos, df_desc = prep.process_sectorial_table(wb[sheet_name], sheet_name)
            data[sheet_name] = df_datos
            df_desc['Tabla'] = sheet_name
            descriptores_all.append(df_desc)
            logger.info(f"  {sheet_name}: {len(df_datos)} filas, {len(df_desc)} descriptores")

    # C5
    if 'C5' in wb.sheetnames:
        data['C5'] = prep.process_c5(wb['C5'])
        logger.info(f"  C5: {len(data['C5'])} filas")

    # Combinar descriptores
    if descriptores_all:
        data['descriptores_CIIU'] = pd.concat(descriptores_all, ignore_index=True)

    return data


def ingest_to_db(data, engine, dry_run=False):
    """Inserta datos en PostgreSQL."""
    if dry_run:
        logger.info("=== DRY RUN - No se escribira en la base de datos ===")

    # Validacion previa
    errors = []
    for key in ['C1.1', 'C3']:
        if key not in data:
            errors.append(f"Dataset {key} requerido no encontrado")
    if errors:
        for e in errors:
            logger.error(e)
        return False

    # Recolectar periodos unicos
    all_periods = set()
    for key, df in data.items():
        if 'Per\u00edodo' in df.columns:
            all_periods.update(df['Per\u00edodo'].dropna().unique())

    logger.info(f"Periodos unicos: {len(all_periods)}")

    if dry_run:
        logger.info(f"Datasets: {list(data.keys())}")
        for key, df in data.items():
            logger.info(f"  {key}: {len(df)} filas, columnas: {list(df.columns)}")
        logger.info("=== DRY RUN completado sin errores ===")
        return True

    # Ejecucion transaccional
    from sqlalchemy import text
    with engine.begin() as conn:
        # Truncate tables
        for table in ['empleo_sectorial', 'empleo_total', 'sectores_ciiu', 'periodos']:
            conn.execute(text(f"TRUNCATE {table} CASCADE"))
        logger.info("Tablas truncadas")

        # Insertar periodos
        periodo_ids = {}
        for period in sorted(all_periods):
            try:
                anio, trimestre, fecha = parse_period_to_date(period)
                result = conn.execute(text(
                    "INSERT INTO periodos (periodo_texto, anio, trimestre, fecha) "
                    "VALUES (:texto, :anio, :trim, :fecha) RETURNING id"
                ), {'texto': period, 'anio': anio, 'trim': trimestre, 'fecha': fecha})
                periodo_ids[period] = result.scalar()
            except Exception as e:
                logger.warning(f"Periodo '{period}': {e}")
        logger.info(f"Periodos insertados: {len(periodo_ids)}")

        # Insertar descriptores
        if 'descriptores_CIIU' in data:
            df_desc = data['descriptores_CIIU']
            count = 0
            for _, row in df_desc.iterrows():
                try:
                    conn.execute(text(
                        "INSERT INTO sectores_ciiu (codigo, descripcion, nivel, tabla_origen) "
                        "VALUES (:cod, :desc, :nivel, :tabla) "
                        "ON CONFLICT (codigo, tabla_origen) DO NOTHING"
                    ), {
                        'cod': str(row['C\u00f3digo']),
                        'desc': str(row['Descripci\u00f3n']),
                        'nivel': str(row.get('Tabla', '')),
                        'tabla': str(row.get('Tabla', ''))
                    })
                    count += 1
                except Exception as e:
                    logger.warning(f"Descriptor: {e}")
            logger.info(f"Descriptores insertados: {count}")

        # C1.1, C1.2: empleo total
        for key, serie in [('C1.1', 'estacional'), ('C1.2', 'desest')]:
            if key not in data:
                continue
            df = data[key]
            count = 0
            for _, row in df.iterrows():
                pid = periodo_ids.get(row.get('Per\u00edodo'))
                if pid is None:
                    continue
                empleo = row.get('Empleo')
                conn.execute(text(
                    "INSERT INTO empleo_total (periodo_id, serie, sector, empleo) "
                    "VALUES (:pid, :serie, NULL, :empleo) "
                    "ON CONFLICT DO NOTHING"
                ), {'pid': pid, 'serie': serie, 'empleo': empleo})
                count += 1
            logger.info(f"{key}: {count} registros insertados")

        # C3, C4, C5, C6, C7: empleo sectorial
        for key in ['C3', 'C4', 'C5', 'C6', 'C7']:
            if key not in data:
                continue
            df = data[key]
            count = 0
            for _, row in df.iterrows():
                pid = periodo_ids.get(row.get('Per\u00edodo'))
                if pid is None:
                    continue
                conn.execute(text(
                    "INSERT INTO empleo_sectorial (periodo_id, tabla_origen, codigo_sector, empleo) "
                    "VALUES (:pid, :tabla, :codigo, :empleo) "
                    "ON CONFLICT DO NOTHING"
                ), {
                    'pid': pid,
                    'tabla': key,
                    'codigo': str(row.get('Sector', '')),
                    'empleo': row.get('Empleo')
                })
                count += 1
            logger.info(f"{key}: {count} registros insertados")

    logger.info("Ingesta completada exitosamente")
    return True


def main():
    parser = argparse.ArgumentParser(description='Cargar datos a PostgreSQL')
    parser.add_argument('excel_path', nargs='?', help='Ruta al archivo Excel')
    parser.add_argument('--from-csv', action='store_true', help='Cargar desde CSV procesados')
    parser.add_argument('--dry-run', action='store_true', help='Validar sin escribir a BD')
    args = parser.parse_args()

    if not args.excel_path and not args.from_csv:
        parser.error("Especifique un archivo Excel o use --from-csv")

    # Cargar datos
    if args.from_csv:
        logger.info("Cargando desde CSV...")
        data = load_csv_data()
    else:
        logger.info(f"Cargando desde Excel: {args.excel_path}")
        if not os.path.exists(args.excel_path):
            logger.error(f"Archivo no encontrado: {args.excel_path}")
            sys.exit(1)
        data = load_excel_data(args.excel_path)

    if not data:
        logger.error("No se pudieron cargar los datos")
        sys.exit(1)

    if args.dry_run:
        ingest_to_db(data, engine=None, dry_run=True)
    else:
        db_url = os.environ.get('DATABASE_URL')
        if not db_url:
            logger.error("DATABASE_URL no configurada. Use --dry-run para validar sin BD.")
            sys.exit(1)

        from sqlalchemy import create_engine
        engine = create_engine(db_url)

        # Crear schema si no existe
        schema_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'schema.sql')
        if os.path.exists(schema_path):
            with open(schema_path, 'r') as f:
                from sqlalchemy import text
                with engine.begin() as conn:
                    for statement in f.read().split(';'):
                        statement = statement.strip()
                        if statement:
                            conn.execute(text(statement))
            logger.info("Schema creado/verificado")

        success = ingest_to_db(data, engine, dry_run=False)
        sys.exit(0 if success else 1)


if __name__ == '__main__':
    main()
