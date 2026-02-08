"""
Script de preprocesamiento optimizado para el Visualizador de Boletines de Empleo
Versión con estructura simplificada y tabla maestra de descriptores
Última actualización: 12 de agosto de 2025
"""

import pandas as pd
import numpy as np
import os
import re
import logging
import openpyxl
from pathlib import Path

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# =====================================================================
# FUNCIONES AUXILIARES GENERALES
# =====================================================================

def is_valid_period(period_str):
    """
    Verifica si un string es un período válido.
    Rechaza comparaciones con "/" y elementos de navegación.
    """
    if not period_str:
        return False
    
    period_str = str(period_str).strip()
    
    # Rechazar si contiene "/" (es una comparación)
    if '/' in period_str:
        return False
    
    # Rechazar si es "Volver al índice" u otro texto de navegación
    if 'volver' in period_str.lower() or 'índice' in period_str.lower():
        return False
    
    # Rechazar si es una nota, fuente o variaciones
    keywords_to_reject = ['nota', 'fuente', 'aclaración', 'variaciones', 'total general']
    if any(keyword in period_str.lower() for keyword in keywords_to_reject):
        return False
    
    # Patrón para períodos válidos
    pattern = r'[1-4](º|°|er|do|er|ro|to)?\s*(trim|Trim)\s*\d{4}'
    
    return bool(re.search(pattern, period_str, re.IGNORECASE))

def normalize_period(period):
    """Normaliza el formato de los períodos para consistencia."""
    if pd.isna(period):
        return period
        
    period = str(period).strip()
    
    # Eliminar espacios extras y normalizar formato
    period = re.sub(r'(\d)(º|°)\s+[Tt]rim\s+(\d)', r'\1º Trim \3', period)
    
    # Reemplazar variaciones
    period = re.sub(r'1er', '1º', period, flags=re.IGNORECASE)
    period = re.sub(r'2do', '2º', period, flags=re.IGNORECASE)
    period = re.sub(r'3(er|ro)', '3º', period, flags=re.IGNORECASE)
    period = re.sub(r'4to', '4º', period, flags=re.IGNORECASE)
    
    # Asegurar formato consistente
    period = re.sub(r'trim', 'Trim', period, flags=re.IGNORECASE)
    
    # Cambiar ° por º para consistencia
    period = period.replace('°', 'º')
    
    # Eliminar espacios extras
    period = re.sub(r'\s+', ' ', period)
    
    return period

# =====================================================================
# PROCESAMIENTO DE SERIES TEMPORALES (C1.1, C1.2, C2.1, C2.2)
# =====================================================================

def process_time_series(ws, sheet_name):
    """
    Procesa hojas de series temporales (C1.1, C1.2, C2.1, C2.2).
    Excluye filas con comparaciones (períodos con '/').
    """
    logger.info(f"Procesando serie temporal: {sheet_name}")
    
    # Leer todos los datos
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)
    
    if len(data) < 3:
        logger.error(f"Hoja {sheet_name} tiene muy pocas filas")
        return pd.DataFrame()
    
    # Los encabezados están típicamente en la fila 3 (índice 2)
    header_row_idx = 2
    headers = data[header_row_idx]
    
    # Determinar estructura según la hoja
    if sheet_name in ['C1.1', 'C1.2']:
        # Series simples con Total
        if sheet_name == 'C1.1':
            columns = ['Período', 'Empleo', 'Var. % interanual']
        else:  # C1.2
            columns = ['Período', 'Empleo', 'Var. % trimestral', 'Var. % interanual']
    else:  # C2.1, C2.2
        # Series por sector
        columns = [
            'Período',
            'Agricultura, ganadería y pesca',
            'Minería y petróleo (3)',
            'Industria',
            'Electricidad, gas y agua (3)',
            'Construcción',
            'Comercio',
            'Servicios',
            'Total'
        ]
    
    # Procesar datos
    datos_lista = []
    filas_excluidas = 0
    
    for row_idx in range(header_row_idx + 1, len(data)):
        row = data[row_idx]
        
        if not row or row[0] is None:
            continue
        
        periodo = str(row[0]).strip()
        
        # Validar período
        if not is_valid_period(periodo):
            logger.debug(f"{sheet_name}: Excluyendo período inválido: {periodo[:50]}")
            filas_excluidas += 1
            continue
        
        # Crear fila de datos
        fila_datos = {'Período': normalize_period(periodo)}
        
        # Extraer valores según la estructura
        for idx, col_name in enumerate(columns[1:], start=1):
            if idx < len(row) and row[idx] is not None:
                try:
                    fila_datos[col_name] = float(row[idx])
                except (ValueError, TypeError):
                    fila_datos[col_name] = None
            else:
                fila_datos[col_name] = None
        
        datos_lista.append(fila_datos)
    
    df = pd.DataFrame(datos_lista)
    
    if not df.empty:
        logger.info(f"{sheet_name}: {len(df)} filas procesadas, {filas_excluidas} excluidas")
    else:
        logger.warning(f"{sheet_name}: Sin datos después del procesamiento")
    
    return df

# =====================================================================
# PROCESAMIENTO DE TABLAS SECTORIALES (SOLO DATOS, SIN DESCRIPTORES)
# =====================================================================

def process_sectorial_table(ws, sheet_name, codigo_col_idx=0, descripcion_col_idx=1):
    """
    Procesa tablas sectoriales (C3, C4, C6, C7) extrayendo solo los datos.
    Los descriptores se guardan en una tabla maestra separada.
    """
    logger.info(f"Procesando {sheet_name}...")
    
    # Leer datos
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)
    
    # Encabezados en fila 3
    header_row_idx = 2
    headers = data[header_row_idx]
    
    # Filtrar períodos válidos
    periodo_columns = []
    for idx, header in enumerate(headers[2:], start=2):
        if header and is_valid_period(str(header)):
            periodo_columns.append((idx, normalize_period(str(header))))
    
    logger.info(f"{sheet_name}: {len(periodo_columns)} períodos válidos encontrados")
    
    # Procesar datos
    datos_lista = []
    descriptores = {}
    
    for row_idx in range(header_row_idx + 1, len(data)):
        row = data[row_idx]
        
        if not row or row[codigo_col_idx] is None:
            continue
        
        codigo = str(row[codigo_col_idx]).strip()
        descripcion = str(row[descripcion_col_idx]).strip() if row[descripcion_col_idx] else ''
        
        # Validar código según la tabla
        if sheet_name == 'C3':
            # Letra (A-O)
            if not codigo or len(codigo) > 1 or not codigo.isalpha():
                continue
            if codigo.lower() == 'total general':
                continue
        elif sheet_name == 'C4':
            # 2 dígitos
            try:
                codigo_num = int(codigo)
                if codigo_num < 1 or codigo_num > 99:
                    continue
            except ValueError:
                continue
        elif sheet_name == 'C6':
            # 3 dígitos
            try:
                codigo_num = int(codigo)
                if codigo_num < 10 or codigo_num > 999:
                    continue
            except ValueError:
                continue
        elif sheet_name == 'C7':
            # 4 dígitos (aunque algunos son de 3)
            try:
                codigo_num = int(codigo)
                if codigo_num < 100 or codigo_num > 9999:
                    continue
            except ValueError:
                continue
        
        # Guardar descriptor
        descriptores[codigo] = descripcion
        
        # Extraer valores
        for col_idx, periodo in periodo_columns:
            if col_idx < len(row) and row[col_idx] is not None:
                valor = row[col_idx]
                
                # Manejar valores especiales
                if isinstance(valor, str):
                    valor_lower = valor.lower().strip()
                    if valor_lower in ['s.d.', 'sd', 's.d', 'n.d.', '-', '...']:
                        valor_num = None
                    else:
                        try:
                            valor_num = float(valor)
                        except:
                            valor_num = None
                else:
                    try:
                        valor_num = float(valor)
                    except:
                        valor_num = None
                
                datos_lista.append({
                    'Sector': codigo,  # Usar solo el código
                    'Período': periodo,
                    'Empleo': valor_num
                })
    
    # Crear DataFrame de datos
    df_datos = pd.DataFrame(datos_lista)
    
    # Crear DataFrame de descriptores
    df_descriptores = pd.DataFrame(list(descriptores.items()), columns=['Código', 'Descripción'])
    
    if not df_datos.empty:
        logger.info(f"{sheet_name}: {len(descriptores)} descriptores, {len(df_datos)} registros")
    
    return df_datos, df_descriptores

def process_c5(ws):
    """
    Procesa C5: Estructura jerárquica sector/tamaño.
    """
    logger.info("Procesando C5 (Sector/Tamaño)...")
    
    # Leer datos
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)
    
    # Encabezados en fila 3
    header_row_idx = 2
    headers = data[header_row_idx]
    
    # Filtrar períodos válidos
    periodo_columns = []
    for idx, header in enumerate(headers[1:], start=1):
        if header and is_valid_period(str(header)):
            periodo_columns.append((idx, normalize_period(str(header))))
    
    logger.info(f"C5: {len(periodo_columns)} períodos válidos encontrados")
    
    # Definir categorías
    sectores_principales = ['industria', 'comercio', 'servicios', 'total']
    tamaños_empresa = ['grandes', 'medianas', 'pequeñas', 'micro']
    
    # Procesar datos
    datos_lista = []
    sector_actual = None
    
    for row_idx in range(header_row_idx + 1, len(data)):
        row = data[row_idx]
        
        if not row or row[0] is None:
            continue
        
        categoria = str(row[0]).strip().lower()
        
        # Detectar si es sector o tamaño
        es_sector = any(s in categoria for s in sectores_principales)
        es_tamaño = any(t in categoria for t in tamaños_empresa)
        
        if es_sector:
            sector_actual = categoria.title()
            tamaño = 'Total'
        elif es_tamaño and sector_actual:
            tamaño = categoria.title()
        else:
            continue
        
        # Crear identificador único
        if es_sector:
            sector_id = categoria.title()
        else:
            sector_id = f"{sector_actual}_{tamaño}"
        
        # Extraer valores
        for col_idx, periodo in periodo_columns:
            if col_idx < len(row) and row[col_idx] is not None:
                try:
                    valor_num = float(row[col_idx])
                    datos_lista.append({
                        'Sector': sector_id,
                        'Período': periodo,
                        'Empleo': valor_num
                    })
                except (ValueError, TypeError):
                    pass
    
    # Crear DataFrame
    df_datos = pd.DataFrame(datos_lista)
    
    if not df_datos.empty:
        logger.info(f"C5: {len(df_datos)} registros procesados")
    
    return df_datos

# =====================================================================
# FUNCIÓN PARA CREAR TABLA MAESTRA DE DESCRIPTORES
# =====================================================================

def create_master_descriptors(all_descriptors, output_dir):
    """
    Crea una tabla maestra con todos los descriptores CIIU.
    """
    if not all_descriptors:
        logger.warning("No hay descriptores para consolidar")
        return
    
    # Consolidar todos los descriptores
    df_maestro = pd.concat(all_descriptors, ignore_index=True)
    
    # Reorganizar columnas
    df_maestro = df_maestro[['Código', 'Descripción', 'Nivel', 'Tabla']]
    
    # Guardar tabla maestra
    output_file = os.path.join(output_dir, 'descriptores_CIIU.csv')
    df_maestro.to_csv(output_file, index=False, encoding='utf-8-sig')
    
    logger.info(f"Tabla maestra de descriptores creada: {len(df_maestro)} entradas")
    
    return df_maestro

# =====================================================================
# FUNCIÓN PRINCIPAL DE PREPROCESAMIENTO
# =====================================================================

def preprocess_excel(input_file, output_dir='../data/processed'):
    """
    Función principal que procesa todo el archivo Excel.
    Genera solo los archivos necesarios con estructura optimizada.
    """
    logger.info("="*80)
    logger.info("INICIANDO PREPROCESAMIENTO OPTIMIZADO")
    logger.info("="*80)
    
    # Verificar archivo de entrada
    if not os.path.exists(input_file):
        logger.error(f"Archivo no encontrado: {input_file}")
        return
    
    # Crear directorio de salida si no existe
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        logger.info(f"Directorio creado: {output_dir}")
    else:
        logger.info(f"Usando directorio existente: {output_dir}")
    
    # Cargar el archivo Excel
    logger.info(f"Cargando archivo: {input_file}")
    wb = openpyxl.load_workbook(input_file, data_only=True)
    
    # Procesar cada hoja
    sheets_to_process = ['C1.1', 'C1.2', 'C2.1', 'C2.2', 'C3', 'C4', 'C5', 'C6', 'C7']
    results = {}
    all_descriptors = []
    
    for sheet_name in sheets_to_process:
        if sheet_name in wb.sheetnames:
            logger.info(f"\nProcesando hoja: {sheet_name}")
            ws = wb[sheet_name]
            
            try:
                if sheet_name in ['C1.1', 'C1.2', 'C2.1', 'C2.2']:
                    # Series temporales - guardar directamente
                    df = process_time_series(ws, sheet_name)
                    if not df.empty:
                        output_file = os.path.join(output_dir, f'{sheet_name}.csv')
                        df.to_csv(output_file, index=False, encoding='utf-8-sig')
                        results[sheet_name] = len(df)
                        
                elif sheet_name == 'C5':
                    # C5 tiene estructura especial
                    df = process_c5(ws)
                    if not df.empty:
                        output_file = os.path.join(output_dir, f'{sheet_name}.csv')
                        df.to_csv(output_file, index=False, encoding='utf-8-sig')
                        results[sheet_name] = len(df)
                        
                else:  # C3, C4, C6, C7
                    # Tablas sectoriales con descriptores
                    df_datos, df_descriptores = process_sectorial_table(ws, sheet_name)
                    
                    if not df_datos.empty:
                        # Guardar solo datos (sin descriptores)
                        output_file = os.path.join(output_dir, f'{sheet_name}.csv')
                        df_datos.to_csv(output_file, index=False, encoding='utf-8-sig')
                        results[sheet_name] = len(df_datos)
                        
                        # Añadir descriptores a la lista para consolidar
                        if not df_descriptores.empty:
                            # Determinar nivel según la tabla
                            nivel_map = {
                                'C3': 'Letra',
                                'C4': '2 dígitos',
                                'C6': '3 dígitos',
                                'C7': '4 dígitos'
                            }
                            df_descriptores['Nivel'] = nivel_map[sheet_name]
                            df_descriptores['Tabla'] = sheet_name
                            all_descriptors.append(df_descriptores)
                    
            except Exception as e:
                logger.error(f"Error procesando {sheet_name}: {str(e)}")
                results[sheet_name] = 0
        else:
            logger.warning(f"Hoja {sheet_name} no encontrada en el archivo")
            results[sheet_name] = 0
    
    wb.close()
    
    # Crear tabla maestra de descriptores
    if all_descriptors:
        df_maestro = create_master_descriptors(all_descriptors, output_dir)
    
    # Resumen final
    logger.info("\n" + "="*80)
    logger.info("RESUMEN DE PROCESAMIENTO OPTIMIZADO")
    logger.info("="*80)
    
    total_records = 0
    for sheet, count in results.items():
        logger.info(f"{sheet}: {count:,} registros")
        total_records += count
    
    logger.info(f"\nTotal de registros procesados: {total_records:,}")
    
    # Listar archivos generados (solo los necesarios)
    logger.info("\nArchivos generados:")
    expected_files = [
        'C1.1.csv', 'C1.2.csv', 'C2.1.csv', 'C2.2.csv',
        'C3.csv', 'C4.csv', 'C5.csv', 'C6.csv', 'C7.csv',
        'descriptores_CIIU.csv'
    ]
    
    for file in expected_files:
        file_path = os.path.join(output_dir, file)
        if os.path.exists(file_path):
            size = os.path.getsize(file_path) / 1024  # KB
            logger.info(f"  - {file} ({size:.1f} KB)")
    
    logger.info("\n" + "="*80)
    logger.info("ESTRUCTURA OPTIMIZADA:")
    logger.info("  - 9 archivos de datos (C1.1 a C7)")
    logger.info("  - 1 tabla maestra de descriptores CIIU")
    logger.info("  - Total: 10 archivos (vs 18 anteriores)")
    logger.info("="*80)
    
    return results

# =====================================================================
# FUNCIÓN DE LIMPIEZA
# =====================================================================

def clean_old_files(output_dir='datos_limpios'):
    """
    Elimina archivos antiguos que ya no son necesarios.
    """
    files_to_remove = [
        'C3_completo.csv', 'C3_descriptores.csv',
        'C4_completo.csv', 'C4_descriptores.csv',
        'C5_completo.csv',
        'C6_completo.csv', 'C6_descriptores.csv',
        'C7_completo.csv', 'C7_descriptores.csv'
    ]
    
    removed_count = 0
    for file in files_to_remove:
        file_path = os.path.join(output_dir, file)
        if os.path.exists(file_path):
            os.remove(file_path)
            removed_count += 1
            logger.info(f"Eliminado: {file}")
    
    if removed_count > 0:
        logger.info(f"\nTotal de archivos eliminados: {removed_count}")
    
    return removed_count

# =====================================================================
# EJECUCIÓN PRINCIPAL
# =====================================================================

if __name__ == "__main__":
    # Archivo de entrada
    input_file = '../data/raw/nacional_serie_empleo_trimestral_actualizado241312.xlsx'
    
    # Directorio de salida
    output_dir = '../data/processed'
    
    # Limpiar archivos antiguos primero
    logger.info("Limpiando archivos antiguos...")
    clean_old_files(output_dir)
    
    # Ejecutar preprocesamiento
    results = preprocess_excel(input_file, output_dir)
    
    # Verificación adicional
    if results:
        print("\nVerificación de integridad:")
        print("-" * 40)
        
        # Verificar que no hay períodos con "/"
        for file in ['C1.1.csv', 'C1.2.csv', 'C2.1.csv', 'C2.2.csv', 
                     'C3.csv', 'C4.csv', 'C5.csv', 'C6.csv', 'C7.csv']:
            file_path = os.path.join(output_dir, file)
            if os.path.exists(file_path):
                df = pd.read_csv(file_path)
                if 'Período' in df.columns:
                    problemas = df[df['Período'].str.contains('/', na=False)]
                    if len(problemas) > 0:
                        print(f"  [ADVERTENCIA] {file}: {len(problemas)} períodos con '/'")
                    else:
                        print(f"  [OK] {file}: Sin períodos problemáticos")
        
        # Verificar tabla de descriptores
        desc_file = os.path.join(output_dir, 'descriptores_CIIU.csv')
        if os.path.exists(desc_file):
            df_desc = pd.read_csv(desc_file)
            print(f"\n  [OK] descriptores_CIIU.csv: {len(df_desc)} descriptores totales")
            print(f"       - Letra: {len(df_desc[df_desc['Nivel']=='Letra'])} descriptores")
            print(f"       - 2 dígitos: {len(df_desc[df_desc['Nivel']=='2 dígitos'])} descriptores")
            print(f"       - 3 dígitos: {len(df_desc[df_desc['Nivel']=='3 dígitos'])} descriptores")
            print(f"       - 4 dígitos: {len(df_desc[df_desc['Nivel']=='4 dígitos'])} descriptores")